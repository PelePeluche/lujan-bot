import os
import re
from PyPDF2 import PdfReader
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def read_pdf_as_plain_text(pdf_path):
    reader = PdfReader(pdf_path)
    all_text = ""
    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text()
        all_text += text + "\n"
    return all_text


def extract_fields_from_text(text, filename):
    fields = {"Archivo": filename}
    try:
        # Expresión regular modificada para hacer que el CUIL sea opcional
        # Expresión regular modificada para incluir 'TA (Tasa Administrativa)', 'AUT (Automotor)', 'INM (Inmueble)', 'CI (Comercio e Industria)'
        pattern = r"(?P<nombre>.+?)\s+(?:(?P<cuil>\d+)\s+)?(?P<reparticion>Inmueble|Automotor|Comercio e Industria|TA \(Tasa Administrativa\)|AUT \(Automotor\)|INM \(Inmueble\)|CI \(Comercio e Industria\))\s+(?P<orden>\d+/\d{4})\s+(?P<identificador>\S+)"
        match = re.search(pattern, text)
        if match:
            fields["Nombre o Razón Social"] = match.group("nombre").strip()
            fields["CUIT"] = (
                match.group("cuil").strip()
                if match.group("cuil")
                else "No especificado"
            )
            fields["Repartición"] = match.group("reparticion").strip()
            orden_anio = match.group("orden").split("/")
            fields["Orden"] = orden_anio[0].strip()
            fields["Año"] = orden_anio[1].strip()
            fields["Identificador"] = match.group("identificador").strip()
        else:
            print("No se encontró coincidencia para los campos principales.")
            fields["Nombre o Razón Social"] = None
            fields["CUIT"] = "No especificado"
            fields["Repartición"] = None
            fields["Orden"] = None
            fields["Año"] = None
            fields["Identificador"] = None

        # Extraer 'Domicilio' y 'Matrícula' según la repartición
        lines = text.splitlines()
        for line in lines:
            if re.search(
                r"Nombre o Razón Social CUIT Repartición Orden Identiﬁcador", line
            ):
                index = lines.index(line)
                if index + 1 < len(lines):
                    domicilio_line = lines[index + 1]
                    if fields["Repartición"] == "Inmueble":
                        domicilio_matricula_pattern = r"(.+?)\s+(\d+)$"
                        domicilio_matricula_match = re.search(
                            domicilio_matricula_pattern, domicilio_line
                        )
                        if domicilio_matricula_match:
                            fields["Domicilio"] = domicilio_matricula_match.group(
                                1
                            ).strip()
                            fields["Matrícula"] = domicilio_matricula_match.group(
                                2
                            ).strip()
                        else:
                            fields["Domicilio"] = "No especificado"
                            fields["Matrícula"] = "No especificado"
                    else:
                        fields["Domicilio"] = domicilio_line.strip()
                        fields["Matrícula"] = "No especificado"
                    break

        # Extract 'Monto total'
        monto_total_pattern = r"LUGAR Y FECHA DE LIQUIDACIÓN DEUDA TOTAL\s+Municipalidad de Córdoba\s+\d{1,2}/\d{1,2}/\d{4}\s+(\d+[.,]?\d*)"
        monto_total_match = re.search(monto_total_pattern, text)
        if not monto_total_match:
            monto_total_pattern_alt = (
                r"DEUDA[\s\S]*?TOTAL[\s\S]*?(\d{1,3}(?:\.\d{3})*,\d{2})"
            )
            monto_total_match = re.search(monto_total_pattern_alt, text)
        fields["Monto total"] = (
            monto_total_match.group(1) if monto_total_match else "No especificado"
        )

        # Determinar si es persona física o jurídica
        if fields["CUIT"] and fields["CUIT"].startswith("3"):
            fields["Tipo de Persona"] = "Jurídica"
        else:
            fields["Tipo de Persona"] = "Física"

    except AttributeError:
        print("Error: Could not extract some fields from the text.")
    return fields


def read_pdfs_in_folder(folder_path, output_excel_path):
    data = []
    # Iterate over all PDF files in the given folder and subfolders
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            if filename.endswith(".pdf"):
                pdf_path = os.path.join(root, filename)
                print(f"Reading content from {filename}...")
                text = read_pdf_as_plain_text(pdf_path)
                # print(text)
                fields = extract_fields_from_text(text, filename)
                if fields:
                    print(
                        f"Fields extracted from {filename}: {fields}"
                    )  # Print extracted fields for verification
                    data.append(fields)

    # Convert the data to a DataFrame and save to Excel
    df = pd.DataFrame(data)
    print(df)  # Print the DataFrame to verify the extracted data
    df.to_excel(output_excel_path, index=False)

    # Apply formatting to the Excel file
    workbook = load_workbook(output_excel_path)
    worksheet = workbook.active

    # Apply styles to header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Adjust column widths
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the formatted Excel file
    workbook.save(output_excel_path)
    print(f"Data successfully saved to {output_excel_path}")


if __name__ == "__main__":
    # Example usage
    # Scan only the specific folder requested by the user
    folder_path = "/Users/admin/Documents/Ramiro Bots/lujan-bot/titles/lujan (13-01-2026)"
    output_excel_path = "/Users/admin/Documents/Ramiro Bots/lujan-bot/titles/lujan (13-01-2026).xlsx"
    read_pdfs_in_folder(folder_path, output_excel_path)
    print("Conversion completed!")
