import os
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import planillas_foldername  # si lo usas

# Solo hasta expediente. Todo lo que siga, lo ignoramos.
line_pattern = re.compile(
    r'^(\d+)\s+'                   # (1) Nro de fila
    r'(\d{1,2}/\d{1,2}/\d{4})\s+'  # (2) Fecha dd/mm/yyyy
    r'(\d+)(INM|AUT|CI)\s+'        # (3) Planilla + (4) Tasa
    r'(\d+/\d{4})\s+'              # (5) N° de Orden (ej. 47661/2022)
    r'(\d+)'                       # (6) N° Expediente
)

def parse_line(line: str, pdf_filename: str):
    match = line_pattern.search(line)
    if not match:
        return None
    
    return {
        "Planilla/PDF": pdf_filename,
        "Nro de fila": match.group(1),
        "Fecha Dda.": match.group(2),
        "Planilla": match.group(3),
        "Tasa": match.group(4),
        "N° de Orden": match.group(5),
        "N° Expediente": match.group(6)
    }

def extract_rows_from_pdf(pdf_path):
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            print(f"\n--- Página {page_num} del PDF {pdf_path} ---")
            print(text if text else "Página sin texto")

            if not text:
                continue
            
            for line in text.split('\n'):
                parsed = parse_line(line, os.path.basename(pdf_path))
                if parsed:
                    rows.append(parsed)
    
    return rows

def process_pdfs_in_folder(folder_path, output_excel_path):
    all_data = []
    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]

    if not pdf_files:
        print(f"No se encontraron archivos PDF en {folder_path}")
        return

    for filename in pdf_files:
        pdf_path = os.path.join(folder_path, filename)
        print(f"\n=== Procesando: {filename} ===")
        rows = extract_rows_from_pdf(pdf_path)
        if rows:
            print(f"   Encontradas {len(rows)} filas válidas en {filename}")
            all_data.extend(rows)
        else:
            print(f"   No se encontraron filas que coincidan con la regex en {filename}")

    if not all_data:
        print("No se extrajo ningún dato de los PDFs.")
        return

    df = pd.DataFrame(all_data)
    df.to_excel(output_excel_path, index=False)
    format_excel(output_excel_path)
    print(f"\nDatos guardados exitosamente en '{output_excel_path}'.")

def format_excel(file_path):
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    
    # Formatea la primera fila (encabezados)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
    
    # Ajusta el ancho de columna según el contenido
    for column in worksheet.columns:
        column_values = [cell.value for cell in column if cell.value]
        if column_values:
            max_length = max(len(str(value)) for value in column_values)
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    workbook.save(file_path)

if __name__ == "__main__":
    folder_path = os.path.join("./planillas", planillas_foldername)
    output_excel_path = "output.xlsx"
    process_pdfs_in_folder(folder_path, output_excel_path)
